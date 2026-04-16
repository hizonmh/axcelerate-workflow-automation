import csv
import re
import io
import os
from datetime import datetime
import openpyxl
import pdfplumber
from reconciler import reconcile_transaction, PAYMENT_FROM_AGENTS, TRANSFER_FROM_AGENTS

# Bank account → instance mapping for CSV files (filename-derived account names).
# Accounts not listed here default to "MAC".
BANK_ACCOUNT_INSTANCE: dict[str, str] = {
    "GC Cheque": "NECGC",
    "GC Prepaid": "NECGC",
    "Melbourne Cheque": "NEC",
    "Melbourne Prepaid": "NEC",
}

# Xero Excel account mapping: Row 4 Col A value → (bank_account, instance).
# MAC accounts use the legacy first-word-before-dash extraction and are not listed here.
XERO_ACCOUNT_MAP: dict[str, tuple[str, str]] = {
    "ANZ (Acc - 220788848)": ("GC Cheque", "NECGC"),
    "ANZ Prepaid (Acc - 303727821)": ("GC Prepaid", "NECGC"),
    "ANZ - Business Chq Acc 5996": ("Melbourne Cheque", "NEC"),
    "ANZ Prepaid Brisbane 4743": ("Melbourne Prepaid", "NEC"),
}


def _normalize_payer(text: str) -> str:
    """Normalize payer name for dedup: first 2 words, lowercase, no digits."""
    if not text or text == "#NAME?":
        return ""
    text = text.lower().strip()
    # Remove common prefixes
    text = re.sub(r"^payment\s+from\s+", "", text)
    text = re.sub(r"^transfer\s+from\s+", "", text)
    # Collapse whitespace
    text = re.sub(r"\s+", " ", text).strip()
    # Truncate at the first digit cluster (student IDs, account numbers, references)
    text = re.sub(r"\d.*$", "", text).strip()
    # Remove trailing/leading punctuation
    text = text.strip("- .,()[]")
    # Take only first 2 words — enough to identify the payer, avoids mismatches
    # from Xero appending extra name/reference text after whitespace padding
    words = text.split()[:2]
    return " ".join(words)


def _extract_payer_from_description(description: str) -> str:
    """Extract payer name from bank description when payer field is missing."""
    text = description.strip()
    # Remove "PAYMENT FROM" / "TRANSFER FROM" prefix to get the name
    text = re.sub(r"^PAYMENT\s+FROM\s+", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^TRANSFER\s+FROM\s+", "", text, flags=re.IGNORECASE)
    # Return the full remaining text — _normalize_payer() handles whitespace
    # collapsing, digit truncation, and word limiting consistently for both
    # Bank CSV and Xero paths.
    return text


def _make_dedup_key(date_str: str, amount: float, payer_raw: str, bank_account: str = "") -> str:
    """Create dedup key from date, amount, normalized payer, and bank account.

    Bank account is always included so identical transactions across different
    accounts are never deduped against each other.
    Negative amounts (outgoing: VISA, transfers, multi-pay) use date+amount only
    because payee naming differs between CSV and Xero sources.
    Positive amounts with no meaningful payer also use date+amount only.
    """
    acct = bank_account.lower().strip()
    if amount < 0:
        return f"{acct}|{date_str}|{amount}"
    norm = _normalize_payer(payer_raw)
    if norm:
        return f"{acct}|{date_str}|{amount}|{norm}"
    return f"{acct}|{date_str}|{amount}"


def _resolve_duplicate_dedup_keys(records: list[dict]) -> list[dict]:
    """Append sequence numbers to duplicate dedup keys within a batch.

    When the same payer sends multiple payments of the same amount on the same
    day (e.g. an agent paying for two different students), the base dedup key
    is identical.  This function detects such collisions and appends |2, |3, …
    to the duplicates so each transaction gets a unique key.  The first
    occurrence keeps its original key for backward compatibility with existing
    DB rows.  Sequence is based on file order (the order records appear in the
    source file), which is stable across re-imports of the same file.
    """
    from collections import defaultdict

    # Group record indices by their dedup_key (preserves file order)
    groups: dict[str, list[int]] = defaultdict(list)
    for idx, rec in enumerate(records):
        groups[rec["dedup_key"]].append(idx)

    for key, indices in groups.items():
        if len(indices) < 2:
            continue
        # First occurrence keeps the original key; subsequent ones get |2, |3, …
        for seq, i in enumerate(indices[1:], start=2):
            records[i]["dedup_key"] = f"{key}|{seq}"

    return records


def _extract_known_agent(text: str) -> str:
    """Return the known agent name if found in text, else empty string.

    Checks both PAYMENT_FROM_AGENTS and TRANSFER_FROM_AGENTS lists so that
    Xero dedup keys use the clean agent name instead of the full description
    (which may include student info that differs from the bank CSV payer column).
    """
    if not text:
        return ""
    lower = text.lower()
    # Strip common prefixes so agent matching works on the core name
    lower = re.sub(r"^(payment|transfer)\s+from\s+", "", lower)
    for agent in PAYMENT_FROM_AGENTS:
        if lower.startswith(agent):
            return agent
    for agent in TRANSFER_FROM_AGENTS:
        if lower.startswith(agent):
            return agent
    return ""


def _bank_account_from_filename(filename: str) -> str:
    """Extract bank account name from filename (e.g. 'Adelaide.csv' → 'Adelaide')."""
    import os
    name = os.path.splitext(os.path.basename(filename))[0]
    return name.strip()


def _is_combined_csv(first_row: list[str]) -> bool:
    """Detect if CSV is the combined multi-bank format (has Source.Name header)."""
    if not first_row:
        return False
    return first_row[0].strip().lower() in ("source.name", "source name")


def parse_combined_bank_csv(file_content: bytes | str) -> list[dict]:
    """Parse combined multi-bank CSV (Source.Name, Column1..Column8 with header row).

    This is the primary bank export format — all banks merged into one file.
    Runs reconciliation automatically to populate student and payment_method.
    """
    if isinstance(file_content, bytes):
        file_content = file_content.decode("utf-8-sig")

    records = []
    reader = csv.reader(io.StringIO(file_content))
    rows = list(reader)

    if not rows:
        return []

    # Skip header row
    start_idx = 1 if _is_combined_csv(rows[0]) else 0

    # Process rows, handling multi-line continuation
    i = start_idx
    while i < len(rows):
        row = rows[i]
        if len(row) < 3:
            i += 1
            continue

        source_name = row[0].strip() if len(row) > 0 else ""
        date_raw = row[1].strip() if len(row) > 1 else ""
        amount_raw = row[2].strip() if len(row) > 2 else ""
        col_d = row[3].strip() if len(row) > 3 else ""  # description
        col_e = row[4].strip() if len(row) > 4 else ""  # payer
        col_f = row[5].strip() if len(row) > 5 else ""  # payee/entity
        # col_g (row[6]) is usually empty — skipped
        col_h = row[7].strip() if len(row) > 7 else ""  # reference 1
        col_i = row[8].strip() if len(row) > 8 else ""  # reference 2

        # Parse date
        try:
            date_obj = datetime.strptime(date_raw, "%d/%m/%Y")
            date_str = date_obj.strftime("%Y-%m-%d")
        except ValueError:
            # Non-date row = continuation text for previous row's Col I
            if records and date_raw:
                records[-1]["payment_note"] = (records[-1]["payment_note"] + " " + date_raw).strip()
            i += 1
            continue

        # Parse amount
        try:
            amount = float(amount_raw.replace(",", ""))
        except ValueError:
            i += 1
            continue

        # Bank account from Source.Name (e.g. "Adelaide.csv" → "Adelaide")
        bank_account = re.sub(r"\.csv$", "", source_name, flags=re.IGNORECASE).strip()
        instance = BANK_ACCOUNT_INSTANCE.get(bank_account, "MAC")

        # Determine payer for dedup
        if col_e and col_e != "#NAME?":
            dedup_source = col_e
        elif col_e == "#NAME?" or col_d.upper().startswith(("PAYMENT FROM", "TRANSFER FROM")):
            dedup_source = _extract_payer_from_description(col_d)
        else:
            dedup_source = ""

        # Run reconciliation
        recon = reconcile_transaction(
            description=col_d,
            payer=col_e,
            payee=col_f,
            col_h=col_h,
            col_i=col_i,
            amount=amount,
        )

        records.append({
            "date": date_str,
            "amount": amount,
            "description": col_d,
            "payer_name": col_e if col_e != "#NAME?" else "",
            "reference": col_h,
            "payment_note": col_i,
            "source": "Bank CSV",
            "bank_account": bank_account,
            "instance": instance,
            "student": recon["student"],
            "payment_method": recon["payment_method"],
            "dedup_key": _make_dedup_key(date_str, amount, dedup_source, bank_account),
        })

        i += 1

    return _resolve_duplicate_dedup_keys(records)


def parse_bank_csv(file_content: bytes | str, bank_account: str = "") -> list[dict]:
    """Parse single-bank Adelaide-style CSV (no header row, 8 columns).

    Columns: Date, Amount, Full Description, Payer Name, Payee Name, (empty), Reference, Payment Note
    """
    if isinstance(file_content, bytes):
        file_content = file_content.decode("utf-8-sig")

    # Check if this is actually a combined CSV
    first_line = file_content.split("\n", 1)[0]
    first_row = next(csv.reader(io.StringIO(first_line)))
    if _is_combined_csv(first_row):
        return parse_combined_bank_csv(file_content)

    records = []
    reader = csv.reader(io.StringIO(file_content))

    for row in reader:
        if len(row) < 2:
            continue

        # Parse date — format is DD/MM/YYYY
        date_raw = row[0].strip()
        try:
            date_obj = datetime.strptime(date_raw, "%d/%m/%Y")
            date_str = date_obj.strftime("%Y-%m-%d")
        except ValueError:
            continue  # Skip non-date rows

        # Parse amount
        try:
            amount = float(row[1].strip().replace(",", ""))
        except ValueError:
            continue

        description = row[2].strip() if len(row) > 2 else ""
        payer_name = row[3].strip() if len(row) > 3 else ""
        payee_name = row[4].strip() if len(row) > 4 else ""
        # Column 5 = empty
        reference = row[6].strip() if len(row) > 6 else ""
        payment_note = row[7].strip() if len(row) > 7 else ""

        # For dedup
        if payer_name and payer_name != "#NAME?":
            dedup_source = payer_name
        elif payer_name == "#NAME?" or description.upper().startswith(("PAYMENT FROM", "TRANSFER FROM")):
            dedup_source = _extract_payer_from_description(description)
        else:
            dedup_source = ""

        # Run reconciliation
        recon = reconcile_transaction(
            description=description,
            payer=payer_name,
            payee=payee_name,
            col_h=reference,
            col_i=payment_note,
            amount=amount,
        )

        instance = BANK_ACCOUNT_INSTANCE.get(bank_account, "MAC")

        records.append({
            "date": date_str,
            "amount": amount,
            "description": description,
            "payer_name": payer_name if payer_name != "#NAME?" else "",
            "reference": reference,
            "payment_note": payment_note,
            "source": "Bank CSV",
            "bank_account": bank_account,
            "instance": instance,
            "student": recon["student"],
            "payment_method": recon["payment_method"],
            "dedup_key": _make_dedup_key(date_str, amount, dedup_source, bank_account),
        })

    return _resolve_duplicate_dedup_keys(records)


def parse_xero_excel(file_content: bytes) -> list[dict]:
    """Parse Xero Bank Reconciliation Excel file, 'Bank Statement' tab.

    Header at row 6: Date, Description, Date Imported, Reference, Reconciled, Source, Amount, Balance
    Row 4 contains the bank account name (e.g. "Adelaide - 316307842").
    Data starts after 'Statement Lines' label (row 11+).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

    if "Bank Statement" not in wb.sheetnames:
        raise ValueError("Sheet 'Bank Statement' not found. Available: " + ", ".join(wb.sheetnames))

    ws = wb["Bank Statement"]

    # Extract bank account name from row 4 (e.g. "Adelaide - 316307842" → "Adelaide")
    row4_val = ws.cell(4, 1).value
    bank_account = ""
    instance = "MAC"
    if row4_val:
        raw = str(row4_val).strip()
        if raw in XERO_ACCOUNT_MAP:
            bank_account, instance = XERO_ACCOUNT_MAP[raw]
        else:
            bank_account = raw.split(" - ")[0].strip()

    records = []
    data_started = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        # Look for "Statement Lines" marker to start parsing data
        if not data_started:
            if row[0] and str(row[0]).strip() == "Statement Lines":
                data_started = True
            continue

        # Skip empty rows, summary rows, and "Closing Balance" row
        date_val = row[0]
        amount_val = row[6] if len(row) > 6 else None

        if date_val is None or amount_val is None:
            continue

        # Stop at "Closing Balance" — can appear in col A or col B
        cell_a = str(row[0]).strip().lower() if row[0] else ""
        cell_b = str(row[1]).strip().lower() if row[1] else ""
        if "closing balance" in cell_a or "closing balance" in cell_b:
            break

        # Skip zero-amount summary rows (e.g. balance carry-forward)
        try:
            if float(amount_val) == 0.0 and not cell_b:
                continue
        except (ValueError, TypeError):
            pass

        # Parse date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            try:
                date_obj = datetime.strptime(str(date_val).strip(), "%Y-%m-%d")
                date_str = date_obj.strftime("%Y-%m-%d")
            except ValueError:
                # Could be "Closing Balance" or other text — stop parsing
                break

        # Parse amount
        try:
            amount = float(amount_val)
        except (ValueError, TypeError):
            continue

        description_raw = str(row[1]).strip() if row[1] else ""
        reference_raw = str(row[3]).strip() if row[3] else ""

        # Extract payer name (col B) and reference from col D
        payer_name = description_raw
        # Reference field often has "PAYMENT <note>" — extract the note
        payment_note = ""
        if reference_raw.upper().startswith("PAYMENT"):
            payment_note = reference_raw[len("PAYMENT"):].strip()
        elif reference_raw:
            payment_note = reference_raw

        # Try to extract student ID from description (digits of 7-8 chars)
        ref_match = re.search(r"\b(\d{7,8})\b", description_raw)
        reference = ref_match.group(1) if ref_match else ""

        # Clean payer for reconciliation: strip trailing student IDs/numbers
        payer_clean = re.sub(r"\s+\d{7,8}\b.*$", "", description_raw).strip()

        # Xero descriptions may embed "AGENT_NAME   STUDENT_INFO" in one field.
        # Extract student hint from text after a known agent name.
        student_hint = ""
        desc_lower = description_raw.lower()
        for agent_name in PAYMENT_FROM_AGENTS:
            idx = desc_lower.find(agent_name)
            if idx >= 0:
                after = description_raw[idx + len(agent_name):].strip()
                if after:
                    student_hint = after
                break

        # Combine description + payment note so FUNDS TFER etc. are visible
        full_description = f"{description_raw} {payment_note}".strip()

        # Run reconciliation — map Xero fields to reconciler inputs:
        #   description = combined Col B + Col D text
        #   payer = cleaned payer name (without embedded IDs)
        #   payee = "" (Xero doesn't have Col F)
        #   col_h = student ID if found, else student hint from agent-embedded description
        #   col_i = payment note from Col D (after "PAYMENT" prefix)
        recon = reconcile_transaction(
            description=full_description,
            payer=payer_clean,
            payee="",
            col_h=reference if reference else student_hint,
            col_i=payment_note,
            amount=amount,
        )

        # For dedup, pass the raw description directly to _make_dedup_key.
        # _normalize_payer (called inside _make_dedup_key) handles everything:
        # collapses whitespace, strips prefixes, truncates at first digit, and
        # takes only the first 2 words — producing the same key as Bank CSV.
        dedup_payer = description_raw

        records.append({
            "date": date_str,
            "amount": amount,
            "description": f"{description_raw} | {reference_raw}".strip(" |"),
            "payer_name": payer_name,
            "reference": reference,
            "payment_note": payment_note,
            "source": "Xero",
            "bank_account": bank_account,
            "instance": instance,
            "student": recon["student"],
            "payment_method": recon["payment_method"],
            "dedup_key": _make_dedup_key(date_str, amount, dedup_payer, bank_account),
        })

    wb.close()
    return _resolve_duplicate_dedup_keys(records)


def _extract_ezidebit_location(pdf_text: str) -> str:
    """Extract campus location from PDF text (e.g. 'Macallan College - Brisbane' → 'Brisbane')."""
    match = re.search(r"Macallan\s+College\s*-\s*(\w+)", pdf_text, re.IGNORECASE)
    if match:
        return match.group(1)
    return ""


def parse_ezidebit_pdf(file_content: bytes) -> list[dict]:
    """Parse Ezidebit Processed Payments Report PDF.

    Uses line-based regex over the extracted text because pdfplumber's table
    extraction is fragile across Ezidebit's layout variants (columns merge or
    split depending on how many data rows there are). A settled Paid row
    always has the structure:

      TRANS_DATE SETTLEMENT_DATE PAYER_ID PAYER_NAME... CLIENT_REF Paid $AMT $FEES $CLEARED

    Pending Paid rows (no settlement date yet) and Failed rows are skipped.
    """
    records = []

    # DD/MM/YYYY   DD/MM/YYYY   NNN-NNN-NNN   <anything>   Paid   $amt   $fees   $cleared
    line_re = re.compile(
        r"(?P<trans>\d{2}/\d{2}/\d{4})\s+"
        r"(?P<settle>\d{2}/\d{2}/\d{4})\s+"
        r"(?P<payer_id>\d{3}-\d{3}-\d{3})\s+"
        r"(?P<middle>.+?)\s+"
        r"Paid\s+"
        r"\$(?P<amount>[0-9,]+\.\d{2})\s+"
        r"\$[0-9,]+\.\d{2}\s+"
        r"\$[0-9,]+\.\d{2}"
    )
    id_re = re.compile(r"(MAC\s?\d+|\b\d{7,8}\b)", re.IGNORECASE)

    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
        location = _extract_ezidebit_location(pdf.pages[0].extract_text() or "") if pdf.pages else ""

        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                m = line_re.search(line)
                if not m:
                    continue

                try:
                    date_obj = datetime.strptime(m.group("settle"), "%d/%m/%Y")
                    date_str = date_obj.strftime("%Y-%m-%d")
                except ValueError:
                    continue

                try:
                    amount = float(m.group("amount").replace(",", ""))
                except ValueError:
                    continue

                middle = m.group("middle").strip()
                id_match = id_re.search(middle)
                if not id_match:
                    continue
                student = id_match.group(1).replace(" ", "").upper()
                # Payer name is everything in "middle" before the student ID,
                # minus trailing commas and whitespace.
                payer = middle[:id_match.start()].strip().rstrip(",").strip()

                records.append({
                    "date": date_str,
                    "amount": amount,
                    "description": f"Ezidebit Direct Debit - {payer}" if payer else "Ezidebit Direct Debit",
                    "payer_name": payer,
                    "reference": student,
                    "payment_note": "",
                    "source": "Ezidebit",
                    "bank_account": "EZIDEBIT",
                    "instance": "EZIDEBIT",
                    "location": location,
                    "student": student,
                    "payment_method": "Direct Debit",
                    "status": "OK to Upload",
                    "dedup_key": f"ezidebit|{date_str}|{amount}|{student}",
                })

    return _resolve_duplicate_dedup_keys(records)


def detect_and_parse(filename: str, content: bytes) -> list[dict]:
    """Auto-detect file type and parse."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        bank_account = _bank_account_from_filename(filename)
        return parse_bank_csv(content, bank_account=bank_account)
    elif lower.endswith(".xlsx"):
        return parse_xero_excel(content)
    elif lower.endswith(".pdf"):
        return parse_ezidebit_pdf(content)
    else:
        raise ValueError(f"Unsupported file type: {filename}. Use .csv, .xlsx, or .pdf")
